from flask import Flask, request, render_template
from openpyxl import load_workbook
from datetime import datetime
import math

app = Flask(__name__)

# ---------------- Configuration ----------------
SCALE_XLSX = "salary_scale.xlsx"

MONTHS = [
    "Mar-2025","Apr-2025","May-2025","Jun-2025",
    "Jul-2025","Aug-2025","Sep-2025","Oct-2025",
    "Nov-2025","Dec-2025","Jan-2026","Feb-2026"
]

DA_RATES = {"Mar-2025":0.1275, "Jul-2025":0.1475}
HRA_RATES = {"A":0.20,"B":0.15,"C":0.075}
CCA_AMOUNT = 600
MEDICAL_AMOUNT = 500

TAX_SLABS = [
    (0,400000,0.0),(400000,800000,0.05),(800000,1200000,0.10),
    (1200000,1600000,0.15),(1600000,2000000,0.20),
    (2000000,2400000,0.25),(2400000,math.inf,0.30)
]
STANDARD_DEDUCTION = 75000
REBATE_87A = 75000

# ---------------- Helpers ----------------
def parse_scale(scale_str):
    """Parse scale like 29600-725-32500-800-35700-900"""
    parts = scale_str.split('-')
    numbers = [float(p) for p in parts]
    if len(numbers) < 3 or len(numbers) % 2 == 0:
        raise ValueError("Invalid scale format")
    salaries = numbers[0::2]; increments = numbers[1::2]
    ranges = []
    for i in range(len(increments)):
        ranges.append((salaries[i], salaries[i+1], increments[i]))
    return ranges

def load_all_ranges(filename=SCALE_XLSX):
    wb = load_workbook(filename, data_only=True)
    ws = wb.active
    scales = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        scale_str = str(row[0]).strip()
        if scale_str:
            scales.append(scale_str)
    return [parse_scale(s.strip()) for s in scales], scales

def calculate_increment_from_scales(salary, all_ranges):
    for row_idx, grade_ranges in enumerate(all_ranges, start=1):
        for start, end, increment in grade_ranges:
            if start <= salary < end:
                return salary + increment, row_idx
    return salary, None

def get_da_rate_for_month(month_str):
    month_dt = datetime.strptime(month_str, "%b-%Y")
    items = sorted(
        [(datetime.strptime(k, "%b-%Y"), v) for k,v in DA_RATES.items()],
        key=lambda x:x[0]
    )
    applicable = 0.0
    for dt, rate in items:
        if month_dt >= dt: applicable = rate
        else: break
    return applicable

def compute_income_tax_new_regime(annual_gross):
    taxable = max(0, annual_gross - STANDARD_DEDUCTION)
    tax = 0.0
    for lower, upper, pct in TAX_SLABS:
        if taxable <= lower:
            break
        taxable_slice = min(taxable, upper) - lower
        if taxable_slice > 0:
            tax += taxable_slice * pct

    rebate = 0.0
    if taxable <= 1275000:
        rebate = min(REBATE_87A, tax)

    tax_after_rebate = max(0, tax - rebate)
    cess = tax_after_rebate * 0.04
    return {
        "std_deduction": STANDARD_DEDUCTION,
        "taxable_income": taxable,
        "tax_before_rebate": tax,
        "rebate_applied": rebate,
        "tax_after_rebate": tax_after_rebate,
        "cess": cess,
        "total_tax_liability": tax_after_rebate + cess
    }

# ---------------- Core Processing ----------------
def process_salary_form(data):
    # Personal details
    kgid = data.get("kgid","").strip()
    name = data.get("name","").strip()
    pan = data.get("pan","").strip()
    phone = data.get("phone","").strip()
    address = data.get("address","").strip()
    designation = data.get("designation","").strip()
    group = data.get("group","").strip().upper()
    city_grade = data.get("city_grade","").strip().upper()
    
    # Salary & increments
    basic_salary = float(data.get("basic_salary",0))
    increment_month_raw = data.get("increment","").strip()
    timebond_from = data.get("timebondmonth","")
    leave_encashment = data.get("leave_encashment","NO").strip().upper()
    allowance = float(data.get("allowance",0))

    # Parse months
    def parse_month(m_raw):
        if not m_raw: return None
        m,y = m_raw.split('-',1)
        dt = datetime.strptime(m[:3]+"-"+y,"%b-%Y")
        return dt.strftime("%b-%Y")
    increment_month = parse_month(increment_month_raw)
    timebond_from_month = parse_month(timebond_from)

    # Salary calculation
    all_ranges, scale_strings = load_all_ranges()
    rows = []; current_basic = basic_salary; dec_gross=0

    for month in MONTHS:
        # Apply regular increment
        if increment_month and month == increment_month:
            current_basic,_ = calculate_increment_from_scales(current_basic, all_ranges)
        # Apply timebond increment
        if timebond_from_month and month == timebond_from_month:
            current_basic,_ = calculate_increment_from_scales(current_basic, all_ranges)

        da = round(current_basic * get_da_rate_for_month(month),2)
        hra = round(current_basic * HRA_RATES.get(city_grade,0),2)
        cca = CCA_AMOUNT if city_grade in ("A","B") else 0
        medical = 0 if group in ("A","B") else MEDICAL_AMOUNT
        gross = current_basic + da + hra + cca + medical

        if month=="Dec-2025": dec_gross=gross
        rows.append({"Month":month,"Basic":current_basic,"DA":da,"HRA":hra,"CCA":cca,"Medical":medical,"Gross":gross})

    annual_basic=sum(r["Basic"] for r in rows)
    annual_da=sum(r["DA"] for r in rows)
    annual_hra=sum(r["HRA"] for r in rows)
    annual_cca=sum(r["CCA"] for r in rows)
    annual_medical=sum(r["Medical"] for r in rows)
    annual_gross=sum(r["Gross"] for r in rows)

    # Add encashment + allowance
    el_encashment_amount=0
    if leave_encashment=="YES": el_encashment_amount=0.5*dec_gross
    annual_gross_with_all=annual_gross+el_encashment_amount+allowance

    tax_summary=compute_income_tax_new_regime(annual_gross_with_all)

    return {
        "kgid":kgid,"name":name,"pan":pan,"phone":phone,"address":address,"designation":designation,
        "city_grade":city_grade,"group":group,"increment_month":increment_month,"timebond_from":timebond_from_month,
        "leave_encashment":leave_encashment,"monthly_rows":rows,
        "annual_basic":annual_basic,"annual_da":annual_da,"annual_hra":annual_hra,"annual_cca":annual_cca,
        "annual_medical":annual_medical,"annual_gross":annual_gross,
        "allowance":allowance,"el_encashment_amount":el_encashment_amount,
        "annual_gross_with_all":annual_gross_with_all,"tax_summary":tax_summary
    }

# ---------------- Flask Routes ----------------
@app.route("/calculate_salary",methods=["POST"])
def calculate_salary_route():
    result=process_salary_form(request.form)
    return render_template("report.html", **result)

@app.route("/")
def index():
    return render_template("form.html")

if __name__=="__main__":
    app.run(debug=True,port=5000)
